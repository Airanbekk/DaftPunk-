import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from db_config import engine 

# Библиотеки для экспорта в Excel
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# 1. Настройка путей
charts_path = "charts/"
os.makedirs(charts_path, exist_ok=True)

exports_path = "exports/"
os.makedirs(exports_path, exist_ok=True)


# ----------------------------------------------------------------------
# ЧАСТЬ 1: ВИЗУАЛИЗАЦИИ MATPLOTLIB (60 БАЛЛОВ)
# ----------------------------------------------------------------------

# 1. Круговая диаграмма (Pie Chart) - Распределение пользователей по странам (2 JOIN)
query1 = """
SELECT T3.Name AS country, 
       COUNT(T1.Id) AS user_count
FROM users AS T1
JOIN locations AS T2 ON T1.locationid = T2.id 
JOIN countries AS T3 ON T2.countryid = T3.id 
GROUP BY T3.Name
ORDER BY user_count DESC
LIMIT 8;
"""
df1 = pd.read_sql(query1, engine)
plt.figure(figsize=(8, 8))
df1.set_index("country")["user_count"].plot.pie(
    autopct="%1.1f%%",
    startangle=90,
    pctdistance=0.85, 
    wedgeprops={'edgecolor': 'black'}
)
plt.title("Распределение пользователей по странам (Топ-8)", pad=20)
plt.ylabel("") 
plt.savefig(os.path.join(charts_path, "users_by_country_pie.png"))
plt.close()
print(f"[Pie] {len(df1)} строк | Распределение пользователей по странам")


# 2. Столбчатая диаграмма (Bar Chart) - Топ-10 жанров по числу мероприятий
query2 = """
SELECT T2.Name AS genre, 
       COUNT(T1.Id) AS event_count
FROM events AS T1
JOIN genres AS T2 ON T1.genreid = T2.id 
GROUP BY T2.Name
ORDER BY event_count DESC
LIMIT 10;
"""
df2 = pd.read_sql(query2, engine)
plt.figure(figsize=(10, 6))
plt.bar(df2["genre"], df2["event_count"], color='skyblue')
plt.title("Топ-10 жанров по числу мероприятий")
plt.xlabel("Жанр")
plt.ylabel("Количество мероприятий")
plt.xticks(rotation=45, ha='right')
plt.tight_layout()
plt.savefig(os.path.join(charts_path, "top_genres_bar.png"))
plt.close()
print(f"[Bar] {len(df2)} строк | Топ-10 жанров по числу мероприятий")


# 3. Горизонтальная столбчатая диаграмма (Horizontal Bar Chart) - Топ-10 артистов по числу выступлений
query3 = """
SELECT T1.Name AS artist_name, 
       COUNT(T2.id) AS performances
FROM artists AS T1
JOIN eventartists AS T2 ON T1.id = T2.artistid 
GROUP BY T1.Name
ORDER BY performances DESC
LIMIT 10;
"""
df3 = pd.read_sql(query3, engine)
plt.figure(figsize=(10, 6))
df3 = df3.sort_values(by='performances') 
plt.barh(df3["artist_name"], df3["performances"], color='lightgreen')
plt.title("Топ-10 артистов по числу выступлений")
plt.xlabel("Количество выступлений")
plt.ylabel("Артист")
plt.tight_layout()
plt.savefig(os.path.join(charts_path, "top_artists_barh.png"))
plt.close()
print(f"[Barh] {len(df3)} строк | Топ-10 артистов по выступлениям")


# 4. Линейный график (Line Chart) - Динамика интереса пользователей к мероприятиям по годам (ИСПРАВЛЕНО)
query4 = """
SELECT DATE_PART('year', T1.Date::TIMESTAMP) AS year, 
       COUNT(T2.id) AS total_interest
FROM events AS T1
JOIN eventhistory AS T2 ON T1.id = T2.eventid 
WHERE T2.isinterested::INTEGER = 1
GROUP BY year
ORDER BY year;
"""
df4 = pd.read_sql(query4, engine)
plt.figure(figsize=(10, 6))
plt.plot(df4["year"], df4["total_interest"], marker="o", linestyle='-', color='purple')
plt.title("Динамика интереса пользователей к мероприятиям по годам")
plt.xlabel("Год")
plt.ylabel("Суммарный интерес (IsInterested = 1)")
plt.grid(True, linestyle='--', alpha=0.6)
plt.xticks(df4["year"])
plt.savefig(os.path.join(charts_path, "events_interest_line.png"))
plt.close()
print(f"[Line] {len(df4)} строк | Динамика интереса по годам")


# 5. Гистограмма (Histogram) - Распределение мероприятий по дням недели (ИСПРАВЛЕНО)
query5 = """
SELECT DATE_PART('dow', T1.Date::TIMESTAMP) AS day_of_week
FROM events AS T1;
"""
df5 = pd.read_sql(query5, engine)
plt.figure(figsize=(10, 6))
day_names = ['Вс', 'Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб']
plt.hist(df5["day_of_week"], bins=7, edgecolor="black", color='gold', rwidth=0.85, align='left') 
plt.title("Распределение мероприятий по дням недели")
plt.xlabel("День недели")
plt.ylabel("Количество мероприятий")
plt.xticks(range(7), day_names)
plt.tight_layout()
plt.savefig(os.path.join(charts_path, "events_by_day_histogram.png"))
plt.close()
print(f"[Histogram] {len(df5)} строк | Распределение мероприятий по дням недели")


# 6. Диаграмма рассеяния (Scatter Plot) - Связь средней оценки мероприятия и популярности жанра (ИСПРАВЛЕНО: 2 JOIN)
query6 = """
WITH GenreAvgRate AS (
    SELECT 
        T3.Name AS genre, 
        AVG(T2.Rate::NUMERIC) AS avg_rate, 
        COUNT(T1.Id) AS event_count
    FROM events AS T1
    JOIN eventhistory AS T2 ON T1.id = T2.eventid 
    JOIN genres AS T3 ON T1.genreid = T3.id 
    WHERE T2.Rate IS NOT NULL AND T2.Rate::NUMERIC > 0 
    GROUP BY T3.Name
)
SELECT * FROM GenreAvgRate
WHERE event_count > 1; 
"""
df6 = pd.read_sql(query6, engine)
plt.figure(figsize=(10, 6))
plt.scatter(df6["event_count"], df6["avg_rate"], alpha=0.7, color='red')

for i in range(len(df6)):
    if df6.loc[i, 'event_count'] >= df6['event_count'].quantile(0.75):
        plt.annotate(df6.loc[i, 'genre'], 
                     (df6.loc[i, 'event_count'] + 0.1, df6.loc[i, 'avg_rate']),
                     fontsize=9)

plt.title("Связь популярности жанра (числа мероприятий) и средней оценки")
plt.xlabel("Количество мероприятий в жанре (Популярность)")
plt.ylabel("Средняя оценка мероприятия (отзывы пользователей)")
plt.grid(True, linestyle=':', alpha=0.5)
plt.tight_layout()
plt.savefig(os.path.join(charts_path, "genre_popularity_vs_rating_scatter.png"))
plt.close()
print(f"[Scatter] {len(df6)} строк | Популярность жанра vs. Средняя оценка")






# ----------------------------------------------------------------------
# ЧАСТЬ 2: ВРЕМЕННОЙ ПОЛЗУНОК PLOTLY (15 БАЛЛОВ)
# ----------------------------------------------------------------------



# analytics.py (ПОЛНЫЙ ИСПРАВЛЕННЫЙ БЛОК PLOTLY)

query_plotly = """
SELECT 
       DATE_PART('year', T1.Date::TIMESTAMP) AS year, 
       DATE_PART('month', T1.Date::TIMESTAMP) AS month,
       T3.Name AS country,
       COUNT(T1.Id) AS event_count
FROM events AS T1
JOIN locations AS T2 ON T1.locationid = T2.id     
JOIN countries AS T3 ON T2.countryid = T3.id      
GROUP BY year, month, T3.Name
ORDER BY year, month, T3.Name;
"""
# 1. ЗАГРУЗКА ДАННЫХ
df_plotly = pd.read_sql(query_plotly, engine) 

# 2. ОПРЕДЕЛЕНИЕ СЕЗОНОВ
def get_season(month):
    if month in [12, 1, 2]: return 'Зима'
    if month in [3, 4, 5]: return 'Весна'
    if month in [6, 7, 8]: return 'Лето'
    if month in [9, 10, 11]: return 'Осень'

df_plotly['season'] = df_plotly['month'].apply(get_season)

print(f"\n[Plotly] {len(df_plotly)} строк | Динамика мероприятий по странам и временам года.")

if len(df_plotly['year'].unique()) > 1:
    
    # Агрегация: СУММИРУЕТ мероприятия по году и стране (исправляет две точки у США)
    df_grouped = df_plotly.groupby(['year', 'country']).agg(
        event_count=('event_count', 'sum'),
        seasons=('season', lambda x: ', '.join(sorted(x.unique())))
    ).reset_index()

    # Фиксированный список стран (исправляет "полет" точек)
    country_order = sorted(df_grouped['country'].unique())

    fig = px.scatter(
        df_grouped, # ИСПОЛЬЗУЕМ АГРЕГИРОВАННЫЕ ДАННЫЕ
        x="country", 
        y="event_count",
        animation_frame="year", 
        size="event_count", 
        color="country", 
        hover_data={"seasons": True, "event_count": True, "country": True}, 
        log_y=True, 
        title="Динамика числа мероприятий по странам (агрегировано по сезонам)"
    )

    # 1. ФИКСАЦИЯ ПОРЯДКА ОСИ X (исправляет "полет" точек)
    fig.update_layout(
        xaxis={'categoryorder':'array', 'categoryarray': country_order}
    )

    # 2. ОСТАЛЬНЫЕ НАСТРОЙКИ
    fig.layout.updatemenus[0].buttons[0].args[1]['frame']['duration'] = 1500
    fig.update_layout(
        xaxis_title="Страна",
        yaxis_title="Количество мероприятий (логарифмическая шкала)",
        yaxis_tickformat='.2s'
    )
    
    fig.show()
    print("Создан интерактивный график Plotly: агрегирован по году и стране.")
else:
    print("Недостаточно данных для анимации Plotly (найден только один год).")




# ----------------------------------------------------------------------
# ЧАСТЬ 3: ЭКСПОРТ В EXCEL С ФОРМАТИРОВАНИЕМ (25 БАЛЛОВ)
# ----------------------------------------------------------------------

# ИСХОДНЫЙ, ПОДРОБНЫЙ SQL-ЗАПРОС ДЛЯ ЭКСЕЛЯ
query_excel = """
SELECT 
       DATE_PART('year', T1.Date::TIMESTAMP) AS year, 
       DATE_PART('month', T1.Date::TIMESTAMP) AS month,
       T3.Name AS country,
       T1.Name AS event_name,
       T1.Date AS event_date,
       COUNT(T1.Id) AS event_count -- ЭТУ КОЛОНКУ МЫ БУДЕМ ФОРМАТИРОВАТЬ
FROM events AS T1
JOIN locations AS T2 ON T1.locationid = T2.id     
JOIN countries AS T3 ON T2.countryid = T3.id      
GROUP BY year, month, T3.Name, T1.Name, T1.Date
ORDER BY year DESC, T3.Name, event_count DESC;
"""

df_excel = pd.read_sql(query_excel, engine)

# Путь для сохранения файла
file_path = os.path.join('exports', 'analytical_report.xlsx')

try:
    # Сохраняем DataFrame в Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_excel.to_excel(writer, sheet_name='Аналитика_Мероприятий', index=False)
        
        workbook = writer.book
        sheet = writer.sheets['Аналитика_Мероприятий']
        
        # 1. ТРЕБОВАНИЕ: ЗАКРЕПЛЕНИЕ ЗАГОЛОВКОВ
        sheet.freeze_panes = 'A2' 
        
        # 2. ТРЕБОВАНИЕ: АВТОФИЛЬТРЫ
        max_col_letter = get_column_letter(sheet.max_column)
        sheet.auto_filter.ref = f"A1:{max_col_letter}{sheet.max_row}"

        # 3. ТРЕБОВАНИЕ: УСЛОВНОЕ ФОРМАТИРОВАНИЕ (ЦВЕТОВЫЕ ШКАЛЫ)
        # Колонка 'event_count' (6-й столбец)
        event_count_col = get_column_letter(df_excel.columns.get_loc('event_count') + 1) 
        
        red_color = openpyxl.styles.Color(rgb='FF0000') 
        green_color = openpyxl.styles.Color(rgb='00FF00') 

        rule = ColorScaleRule(
            start_type='min', 
            start_color=red_color,
            end_type='max', 
            end_color=green_color
        )
        
        sheet.conditional_formatting.add(f'{event_count_col}2:{event_count_col}{sheet.max_row}', rule)

    print(f"\n[Excel] Отчет успешно создан и сохранен в: {file_path}")

except Exception as e:
    print(f"\n[Excel] Ошибка при экспорте в Excel: {e}")